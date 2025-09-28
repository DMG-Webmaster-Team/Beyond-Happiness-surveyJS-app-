#!/usr/bin/env python3
"""
Create enhanced Excel file with comprehensive multilingual happiness survey data
Including all UI elements, general terms, and system messages in Arabic
"""

import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

def load_json_data():
    """Load all JSON data files"""
    
    # Load questions data
    with open('data/happiness-questions-multilingual.json', 'r', encoding='utf-8') as f:
        questions_data = json.load(f)
    
    # Load characters data
    with open('data/happiness-characters-multilingual.json', 'r', encoding='utf-8') as f:
        characters_data = json.load(f)
    
    return questions_data, characters_data

def create_questions_sheet(wb, questions_data):
    """Create Questions sheet with all questions and choices"""
    
    ws = wb.create_sheet("Questions", 0)
    
    # Headers
    headers = [
        "Question ID", "Category (EN)", "Category (AR)", "Question (English)", "Question (Arabic)",
        "Choice 1 (EN)", "Choice 1 (AR)", "Value 1",
        "Choice 2 (EN)", "Choice 2 (AR)", "Value 2", 
        "Choice 3 (EN)", "Choice 3 (AR)", "Value 3",
        "Choice 4 (EN)", "Choice 4 (AR)", "Value 4",
        "Choice 5 (EN)", "Choice 5 (AR)", "Value 5"
    ]
    
    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Category translations
    category_translations = {
        "Meaning": "المعنى",
        "Delight": "البهجة", 
        "Freedom": "الحرية",
        "Engagement": "الانخراط",
        "Vitality": "الحيوية"
    }
    
    # Get choices data
    choices = questions_data.get('choices', [])
    
    # Write question data
    for i, question in enumerate(questions_data['questions'], 2):
        ws.cell(row=i, column=1, value=question['id'])
        ws.cell(row=i, column=2, value=question['category'])
        ws.cell(row=i, column=3, value=category_translations.get(question['category'], question['category']))
        ws.cell(row=i, column=4, value=question['question']['en'])
        ws.cell(row=i, column=5, value=question['question']['ar'])
        
        # Add choice data
        for j, choice in enumerate(choices):
            choice_col_start = 6 + (j * 3)  # Each choice takes 3 columns (EN, AR, Value)
            ws.cell(row=i, column=choice_col_start, value=choice['text']['en'])
            ws.cell(row=i, column=choice_col_start + 1, value=choice['text']['ar'])
            ws.cell(row=i, column=choice_col_start + 2, value=question['values'][j])
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
        ws.column_dimensions[column_letter].width = adjusted_width

def create_categories_sheet(wb):
    """Create Categories sheet with category translations"""
    
    ws = wb.create_sheet("Categories")
    
    # Categories data
    categories_data = [
        {"Category": "Meaning", "English": "Meaning", "Arabic": "المعنى", 
         "Color": "#7E57C2", "Description (EN)": "Your sense of purpose, spiritual connection, and alignment with core values.",
         "Description (AR)": "إحساسك بالهدف، والاتصال الروحي، والتوافق مع القيم الأساسية."},
        {"Category": "Delight", "English": "Delight", "Arabic": "البهجة", 
         "Color": "#FFCA28", "Description (EN)": "Your capacity for joy, playfulness, creativity, and appreciation of beauty.",
         "Description (AR)": "قدرتك على الفرح، واللعب، والإبداع، وتقدير الجمال."},
        {"Category": "Freedom", "English": "Freedom", "Arabic": "الحرية", 
         "Color": "#FFA726", "Description (EN)": "Your sense of autonomy, self-expression, and control over your life direction.",
         "Description (AR)": "إحساسك بالاستقلالية والتعبير عن الذات والسيطرة على مسار حياتك."},
        {"Category": "Engagement", "English": "Engagement", "Arabic": "الانخراط", 
         "Color": "#42A5F5", "Description (EN)": "Your level of involvement, focus, and utilization of skills in meaningful activities.",
         "Description (AR)": "مستوى مشاركتك وتركيزك واستخدام مهاراتك في أنشطة ذات مغزى."},
        {"Category": "Vitality", "English": "Vitality", "Arabic": "الحيوية", 
         "Color": "#66BB6A", "Description (EN)": "Your physical and mental energy, health, resilience, and overall well-being.",
         "Description (AR)": "طاقتك البدنية والعقلية وصحتك ومرونتك ورفاهيتك العامة."}
    ]
    
    # Headers
    headers = ["Category", "English", "Arabic", "Brand Color", "Description (English)", "Description (Arabic)"]
    
    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Write category data
    for i, category in enumerate(categories_data, 2):
        ws.cell(row=i, column=1, value=category['Category'])
        ws.cell(row=i, column=2, value=category['English'])
        ws.cell(row=i, column=3, value=category['Arabic'])
        ws.cell(row=i, column=4, value=category['Color'])
        ws.cell(row=i, column=5, value=category['Description (EN)'])
        ws.cell(row=i, column=6, value=category['Description (AR)'])
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 60)  # Cap at 60 characters
        ws.column_dimensions[column_letter].width = adjusted_width

def create_characters_sheet(wb, characters_data):
    """Create Characters sheet with all character names and descriptions"""
    
    ws = wb.create_sheet("Characters")
    
    # Headers
    headers = ["Character ID", "Match Code", "Name (English)", "Name (Arabic)", 
               "Description (English)", "Description (Arabic)", "Avatar URL"]
    
    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Write character data
    for i, character in enumerate(characters_data['characters'], 2):
        ws.cell(row=i, column=1, value=character['id'])
        ws.cell(row=i, column=2, value=character['match'])
        ws.cell(row=i, column=3, value=character['name']['en'])
        ws.cell(row=i, column=4, value=character['name']['ar'])
        ws.cell(row=i, column=5, value=character['description']['en'])
        ws.cell(row=i, column=6, value=character['description']['ar'])
        ws.cell(row=i, column=7, value=character['avatar_url'])
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 70)  # Cap at 70 characters
        ws.column_dimensions[column_letter].width = adjusted_width

def create_choices_sheet(wb, questions_data):
    """Create Choices sheet with answer choices"""
    
    ws = wb.create_sheet("Answer Choices")
    
    # Headers
    headers = ["Choice Value", "Choice Text (English)", "Choice Text (Arabic)", "Description"]
    
    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Get choices data
    choices = questions_data.get('choices', [])
    
    # Write choice data
    descriptions = [
        "Strongly disagree / Never applies",
        "Disagree / Rarely applies", 
        "Neutral / Sometimes applies",
        "Agree / Often applies",
        "Strongly agree / Always applies"
    ]
    
    for i, choice in enumerate(choices, 2):
        ws.cell(row=i, column=1, value=choice['value'])
        ws.cell(row=i, column=2, value=choice['text']['en'])
        ws.cell(row=i, column=3, value=choice['text']['ar'])
        ws.cell(row=i, column=4, value=descriptions[i-2] if i-2 < len(descriptions) else "")
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

def create_general_translations_sheet(wb):
    """Create General Translations sheet with UI elements and system messages"""
    
    ws = wb.create_sheet("General Translations")
    
    # Headers
    headers = ["Category", "Context", "English", "Arabic", "Usage Notes"]
    
    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # General translations data
    translations_data = [
        # Survey UI Elements
        {"Category": "Survey UI", "Context": "Language Selection", "English": "Choose your language", "Arabic": "اختر لغتك", "Usage": "Language selector dropdown"},
        {"Category": "Survey UI", "Context": "Language Selection", "English": "English", "Arabic": "الإنجليزية", "Usage": "Language option"},
        {"Category": "Survey UI", "Context": "Language Selection", "English": "Arabic", "Arabic": "العربية", "Usage": "Language option"},
        {"Category": "Survey UI", "Context": "Navigation", "English": "Next", "Arabic": "التالي", "Usage": "Next button"},
        {"Category": "Survey UI", "Context": "Navigation", "English": "Previous", "Arabic": "السابق", "Usage": "Previous button"},
        {"Category": "Survey UI", "Context": "Navigation", "English": "Submit", "Arabic": "إرسال", "Usage": "Submit button"},
        {"Category": "Survey UI", "Context": "Navigation", "English": "Start Survey", "Arabic": "ابدأ الاستطلاع", "Usage": "Start button"},
        {"Category": "Survey UI", "Context": "Progress", "English": "Question {0} of {1}", "Arabic": "السؤال {0} من {1}", "Usage": "Progress indicator"},
        {"Category": "Survey UI", "Context": "Progress", "English": "Progress", "Arabic": "التقدم", "Usage": "Progress label"},
        
        # Results Page
        {"Category": "Results", "Context": "Page Title", "English": "Your Happiness Profile", "Arabic": "ملفك الشخصي للسعادة", "Usage": "Main results title"},
        {"Category": "Results", "Context": "Page Subtitle", "English": "Discover your unique character and happiness dimensions", "Arabic": "اكتشف شخصيتك الفريدة وأبعاد السعادة الخاصة بك", "Usage": "Results subtitle"},
        {"Category": "Results", "Context": "Character Announcement", "English": "You are a", "Arabic": "أنت", "Usage": "Character reveal text"},
        {"Category": "Results", "Context": "Section Title", "English": "Your Character Description", "Arabic": "وصف شخصيتك", "Usage": "Character description section"},
        {"Category": "Results", "Context": "Section Title", "English": "Overall Happiness Score", "Arabic": "النتيجة الإجمالية للسعادة", "Usage": "Overall score section"},
        {"Category": "Results", "Context": "Section Title", "English": "Overall Happiness Level", "Arabic": "مستوى السعادة الإجمالي", "Usage": "Overall level description"},
        {"Category": "Results", "Context": "Section Title", "English": "Happiness Dimensions Overview", "Arabic": "نظرة عامة على أبعاد السعادة", "Usage": "Chart section title"},
        {"Category": "Results", "Context": "Section Title", "English": "Detailed Happiness Dimensions", "Arabic": "أبعاد السعادة التفصيلية", "Usage": "Detailed breakdown section"},
        
        # Actions
        {"Category": "Actions", "Context": "Button", "English": "Download PDF", "Arabic": "تحميل ملف PDF", "Usage": "PDF download button"},
        {"Category": "Actions", "Context": "Button", "English": "Download Your Report", "Arabic": "حمل تقريرك", "Usage": "Report download button"},
        {"Category": "Actions", "Context": "Button", "English": "Retake Survey", "Arabic": "إعادة الاستطلاع", "Usage": "Retake button"},
        {"Category": "Actions", "Context": "Button", "English": "Share Results", "Arabic": "شارك النتائج", "Usage": "Share button"},
        {"Category": "Actions", "Context": "Link", "English": "Back to Survey", "Arabic": "العودة للاستطلاع", "Usage": "Navigation link"},
        
        # Status Messages
        {"Category": "Messages", "Context": "Loading", "English": "Loading...", "Arabic": "جارٍ التحميل...", "Usage": "Loading indicator"},
        {"Category": "Messages", "Context": "Success", "English": "Survey completed successfully!", "Arabic": "تم إكمال الاستطلاع بنجاح!", "Usage": "Success message"},
        {"Category": "Messages", "Context": "Error", "English": "An error occurred. Please try again.", "Arabic": "حدث خطأ. يرجى المحاولة مرة أخرى.", "Usage": "Generic error message"},
        {"Category": "Messages", "Context": "Validation", "English": "Please answer all questions", "Arabic": "يرجى الإجابة على جميع الأسئلة", "Usage": "Validation message"},
        {"Category": "Messages", "Context": "Cooldown", "English": "You must wait {0} more day(s) before retaking this survey", "Arabic": "يجب أن تنتظر {0} يوم إضافي قبل إعادة هذا الاستطلاع", "Usage": "Cooldown message"},
        
        # PDF Content
        {"Category": "PDF", "Context": "Footer", "English": "Generated by Mountain View Happiness Survey", "Arabic": "تم إنشاؤه بواسطة استطلاع السعادة من ماونتن فيو", "Usage": "PDF footer text"},
        {"Category": "PDF", "Context": "Brand", "English": "Beyond Happiness", "Arabic": "ما وراء السعادة", "Usage": "Brand name"},
        {"Category": "PDF", "Context": "Score Text", "English": "You scored {0}% in the Happiness Survey", "Arabic": "لقد حصلت على {0}% في مسح السعادة", "Usage": "Score description"},
        
        # Navigation & Menu
        {"Category": "Navigation", "Context": "Menu", "English": "Home", "Arabic": "الرئيسية", "Usage": "Home menu item"},
        {"Category": "Navigation", "Context": "Menu", "English": "Dashboard", "Arabic": "لوحة التحكم", "Usage": "Dashboard menu item"},
        {"Category": "Navigation", "Context": "Menu", "English": "Profile", "Arabic": "الملف الشخصي", "Usage": "Profile menu item"},
        {"Category": "Navigation", "Context": "Menu", "English": "Settings", "Arabic": "الإعدادات", "Usage": "Settings menu item"},
        {"Category": "Navigation", "Context": "Menu", "English": "Logout", "Arabic": "تسجيل الخروج", "Usage": "Logout menu item"},
        
        # Form Elements
        {"Category": "Forms", "Context": "Label", "English": "Name", "Arabic": "الاسم", "Usage": "Name field label"},
        {"Category": "Forms", "Context": "Label", "English": "Email", "Arabic": "البريد الإلكتروني", "Usage": "Email field label"},
        {"Category": "Forms", "Context": "Label", "English": "Phone", "Arabic": "الهاتف", "Usage": "Phone field label"},
        {"Category": "Forms", "Context": "Placeholder", "English": "Enter your name", "Arabic": "أدخل اسمك", "Usage": "Name field placeholder"},
        {"Category": "Forms", "Context": "Placeholder", "English": "Enter your email", "Arabic": "أدخل بريدك الإلكتروني", "Usage": "Email field placeholder"},
        
        # Time & Date
        {"Category": "Time", "Context": "Relative", "English": "Today", "Arabic": "اليوم", "Usage": "Today reference"},
        {"Category": "Time", "Context": "Relative", "English": "Yesterday", "Arabic": "أمس", "Usage": "Yesterday reference"},
        {"Category": "Time", "Context": "Relative", "English": "Tomorrow", "Arabic": "غداً", "Usage": "Tomorrow reference"},
        {"Category": "Time", "Context": "Duration", "English": "minute", "Arabic": "دقيقة", "Usage": "Minute unit"},
        {"Category": "Time", "Context": "Duration", "English": "minutes", "Arabic": "دقائق", "Usage": "Minutes plural"},
        {"Category": "Time", "Context": "Duration", "English": "hour", "Arabic": "ساعة", "Usage": "Hour unit"},
        {"Category": "Time", "Context": "Duration", "English": "hours", "Arabic": "ساعات", "Usage": "Hours plural"},
        {"Category": "Time", "Context": "Duration", "English": "day", "Arabic": "يوم", "Usage": "Day unit"},
        {"Category": "Time", "Context": "Duration", "English": "days", "Arabic": "أيام", "Usage": "Days plural"},
        
        # Common Actions
        {"Category": "Common", "Context": "Action", "English": "Save", "Arabic": "حفظ", "Usage": "Save action"},
        {"Category": "Common", "Context": "Action", "English": "Cancel", "Arabic": "إلغاء", "Usage": "Cancel action"},
        {"Category": "Common", "Context": "Action", "English": "Delete", "Arabic": "حذف", "Usage": "Delete action"},
        {"Category": "Common", "Context": "Action", "English": "Edit", "Arabic": "تعديل", "Usage": "Edit action"},
        {"Category": "Common", "Context": "Action", "English": "View", "Arabic": "عرض", "Usage": "View action"},
        {"Category": "Common", "Context": "Action", "English": "Close", "Arabic": "إغلاق", "Usage": "Close action"},
        {"Category": "Common", "Context": "Status", "English": "Active", "Arabic": "نشط", "Usage": "Active status"},
        {"Category": "Common", "Context": "Status", "English": "Inactive", "Arabic": "غير نشط", "Usage": "Inactive status"},
        {"Category": "Common", "Context": "Status", "English": "Completed", "Arabic": "مكتمل", "Usage": "Completed status"},
        {"Category": "Common", "Context": "Status", "English": "Pending", "Arabic": "معلق", "Usage": "Pending status"},
    ]
    
    # Write translation data
    for i, translation in enumerate(translations_data, 2):
        ws.cell(row=i, column=1, value=translation['Category'])
        ws.cell(row=i, column=2, value=translation['Context'])
        ws.cell(row=i, column=3, value=translation['English'])
        ws.cell(row=i, column=4, value=translation['Arabic'])
        ws.cell(row=i, column=5, value=translation['Usage'])
    
    # Auto-adjust column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 40
    ws.column_dimensions['D'].width = 40
    ws.column_dimensions['E'].width = 30

def create_summary_sheet(wb, questions_data, characters_data):
    """Create Summary sheet with overview statistics"""
    
    ws = wb.create_sheet("Summary", 0)
    
    # Title
    title_cell = ws.cell(row=1, column=1, value="Happiness Survey - Complete Multilingual Content")
    title_cell.font = Font(size=16, bold=True, color="FFFFFF")
    title_cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells('A1:D1')
    
    # Statistics
    stats = [
        ["", "", "", ""],
        ["Content Type", "Count", "Languages", "Notes"],
        ["Survey Questions", len(questions_data['questions']), "English + Arabic", "40 questions across 5 categories"],
        ["Answer Choices", len(questions_data.get('choices', [])), "English + Arabic", "5-point Likert scale"],
        ["Happiness Categories", 5, "English + Arabic", "Meaning, Delight, Freedom, Engagement, Vitality"],
        ["Character Types", len(characters_data['characters']), "English + Arabic", "32 unique personality profiles"],
        ["General UI Terms", "50+", "English + Arabic", "Buttons, labels, messages, navigation"],
        ["", "", "", ""],
        ["Category Breakdown:", "", "", ""],
        ["Meaning Questions", len([q for q in questions_data['questions'] if q['category'] == 'Meaning']), "المعنى", "Questions 1-8"],
        ["Delight Questions", len([q for q in questions_data['questions'] if q['category'] == 'Delight']), "البهجة", "Questions 9-16"],
        ["Freedom Questions", len([q for q in questions_data['questions'] if q['category'] == 'Freedom']), "الحرية", "Questions 17-24"],
        ["Engagement Questions", len([q for q in questions_data['questions'] if q['category'] == 'Engagement']), "الانخراط", "Questions 25-32"],
        ["Vitality Questions", len([q for q in questions_data['questions'] if q['category'] == 'Vitality']), "الحيوية", "Questions 33-40"],
        ["", "", "", ""],
        ["Translation Coverage:", "", "", ""],
        ["Survey Content", "100%", "Complete", "All questions, choices, categories"],
        ["Character Profiles", "100%", "Complete", "All names and descriptions"],
        ["UI Elements", "100%", "Complete", "Buttons, labels, navigation"],
        ["System Messages", "100%", "Complete", "Errors, success, validation"],
        ["PDF Content", "100%", "Complete", "Report generation text"],
    ]
    
    # Write statistics
    for i, row in enumerate(stats, 2):
        for j, value in enumerate(row, 1):
            cell = ws.cell(row=i, column=j, value=value)
            if i == 3:  # Header row
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    
    # Auto-adjust column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 40

def main():
    """Main function to create the enhanced Excel file"""
    
    print("🔄 Loading JSON data...")
    questions_data, characters_data = load_json_data()
    
    print("📊 Creating enhanced Excel workbook...")
    wb = Workbook()
    
    # Remove default sheet
    wb.remove(wb.active)
    
    print("📝 Creating Summary sheet...")
    create_summary_sheet(wb, questions_data, characters_data)
    
    print("❓ Creating Questions sheet with category translations...")
    create_questions_sheet(wb, questions_data)
    
    print("🏷️ Creating Categories sheet...")
    create_categories_sheet(wb)
    
    print("👤 Creating Characters sheet...")
    create_characters_sheet(wb, characters_data)
    
    print("✅ Creating Answer Choices sheet...")
    create_choices_sheet(wb, questions_data)
    
    print("🌍 Creating General Translations sheet...")
    create_general_translations_sheet(wb)
    
    # Save the file
    filename = "Happiness_Survey_Complete_Multilingual_Content.xlsx"
    wb.save(filename)
    
    print(f"✅ Enhanced Excel file created successfully: {filename}")
    print("\n📋 File contains:")
    print("   • Summary - Overview and statistics with Arabic category names")
    print("   • Questions - All 40 questions with category translations")
    print("   • Categories - 5 happiness dimensions with full translations")
    print("   • Characters - 32 personality types with translations")
    print("   • Answer Choices - 5-point scale in both languages")
    print("   • General Translations - 50+ UI elements, messages, and terms")
    print("\n🌍 Complete Arabic coverage for:")
    print("   • All survey content and UI elements")
    print("   • Navigation, buttons, and form labels")
    print("   • Status messages and error handling")
    print("   • PDF content and branding")
    print("   • Time/date references and common actions")

if __name__ == "__main__":
    main()
