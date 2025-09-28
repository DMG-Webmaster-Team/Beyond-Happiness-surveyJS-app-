#!/usr/bin/env python3
"""
Add subtype translations to the existing Excel file
"""

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

def add_subtypes_to_excel():
    """Add subtype translations to the General Translations sheet"""
    
    # Load existing workbook
    wb = load_workbook("Happiness_Survey_Complete_Multilingual_Content.xlsx")
    
    # Get the General Translations sheet
    ws = wb["General Translations"]
    
    # Find the last row with data
    last_row = ws.max_row
    
    # Subtype translations to add
    subtype_translations = [
        {"Category": "PDF Subtypes", "Context": "Category Breakdown", "English": "Type A", "Arabic": "النوع A", "Usage": "First subtype in category breakdown"},
        {"Category": "PDF Subtypes", "Context": "Category Breakdown", "English": "Type B", "Arabic": "النوع B", "Usage": "Second subtype in category breakdown"},
        {"Category": "PDF Subtypes", "Context": "Category Breakdown", "English": "Type C", "Arabic": "النوع C", "Usage": "Third subtype in category breakdown"},
        {"Category": "PDF Subtypes", "Context": "Category Breakdown", "English": "Type D", "Arabic": "النوع D", "Usage": "Fourth subtype in category breakdown"},
        {"Category": "PDF Subtypes", "Context": "Section Description", "English": "Subtype Breakdown", "Arabic": "تفصيل الأنواع الفرعية", "Usage": "Subtype section header"},
        {"Category": "PDF Subtypes", "Context": "Question Mapping", "English": "Questions 1-2", "Arabic": "الأسئلة 1-2", "Usage": "Type A question range"},
        {"Category": "PDF Subtypes", "Context": "Question Mapping", "English": "Questions 3-4", "Arabic": "الأسئلة 3-4", "Usage": "Type B question range"},
        {"Category": "PDF Subtypes", "Context": "Question Mapping", "English": "Questions 5-6", "Arabic": "الأسئلة 5-6", "Usage": "Type C question range"},
        {"Category": "PDF Subtypes", "Context": "Question Mapping", "English": "Questions 7-8", "Arabic": "الأسئلة 7-8", "Usage": "Type D question range"},
    ]
    
    # Add subtype translations
    for i, translation in enumerate(subtype_translations, last_row + 1):
        ws.cell(row=i, column=1, value=translation['Category'])
        ws.cell(row=i, column=2, value=translation['Context'])
        ws.cell(row=i, column=3, value=translation['English'])
        ws.cell(row=i, column=4, value=translation['Arabic'])
        ws.cell(row=i, column=5, value=translation['Usage'])
    
    # Update the Summary sheet with subtype information
    summary_ws = wb["Summary"]
    
    # Find the row with "General UI Terms" and update it
    for row in summary_ws.iter_rows(min_row=2, max_row=summary_ws.max_row):
        if row[0].value == "General UI Terms":
            row[1].value = "60+"  # Updated count
            row[3].value = "Buttons, labels, messages, navigation, subtypes"  # Updated notes
            break
    
    # Add a new row for subtype information
    new_row = summary_ws.max_row + 1
    summary_ws.cell(row=new_row, column=1, value="PDF Subtypes")
    summary_ws.cell(row=new_row, column=2, value="4 per category")
    summary_ws.cell(row=new_row, column=3, value="English + Arabic")
    summary_ws.cell(row=new_row, column=4, value="Type A-D breakdown with progress bars")
    
    # Save the updated workbook
    wb.save("Happiness_Survey_Complete_Multilingual_Content.xlsx")
    
    print("✅ Excel file updated with subtype translations!")
    print("📊 Added 9 new subtype-related translations")
    print("📋 Updated summary with subtype information")

if __name__ == "__main__":
    add_subtypes_to_excel()

